#!/usr/bin/env python3
"""
Generate data/bracket.xlsx from data/picks.json + data/scores.json.
Run manually or via GitHub Actions after picks.json is updated.

Requires: pip install openpyxl
"""

import json
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed.  Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

BASE = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
PICKS_PATH  = os.path.join(BASE, "data", "picks.json")
SCORES_PATH = os.path.join(BASE, "data", "scores.json")
OUT_PATH    = os.path.join(BASE, "data", "bracket.xlsx")

ROUND_NAMES  = ["", "First Round", "Conf. Semifinals", "Conf. Finals", "NBA Finals"]
ROUND_POINTS = [0, 1, 2, 4, 8]
GAMES_BONUS  = 1

SERIES = [
    {"id": "E1v8",  "r": 1, "conf": "East"},
    {"id": "E4v5",  "r": 1, "conf": "East"},
    {"id": "E2v7",  "r": 1, "conf": "East"},
    {"id": "E3v6",  "r": 1, "conf": "East"},
    {"id": "W1v8",  "r": 1, "conf": "West"},
    {"id": "W4v5",  "r": 1, "conf": "West"},
    {"id": "W2v7",  "r": 1, "conf": "West"},
    {"id": "W3v6",  "r": 1, "conf": "West"},
    {"id": "EQ1",   "r": 2, "conf": "East"},
    {"id": "EQ2",   "r": 2, "conf": "East"},
    {"id": "WQ1",   "r": 2, "conf": "West"},
    {"id": "WQ2",   "r": 2, "conf": "West"},
    {"id": "ECF",   "r": 3, "conf": "East"},
    {"id": "WCF",   "r": 3, "conf": "West"},
    {"id": "FINALS","r": 4, "conf": None},
]

# ---- colours ----
HDR_FILL   = PatternFill("solid", fgColor="1A1D27")
HDR_FONT   = Font(color="F97316", bold=True)
GREEN_FILL = PatternFill("solid", fgColor="16532A")
RED_FILL   = PatternFill("solid", fgColor="4B1010")
GREY_FILL  = PatternFill("solid", fgColor="22263A")
WHITE_FONT = Font(color="E8EAF0")
DIM_FONT   = Font(color="7B82A3")
BOLD_WHITE = Font(color="E8EAF0", bold=True)

thin = Side(style="thin", color="2E3352")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def cell(ws, row, col, value="", fill=None, font=None, align="left", bold=False):
    c = ws.cell(row=row, column=col, value=value)
    if fill:  c.fill = fill
    if font:  c.font = font
    elif bold: c.font = Font(color="E8EAF0", bold=True)
    else:      c.font = WHITE_FONT
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    c.border = BORDER
    return c


def auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(min_w, min(length + 2, max_w))


def main():
    # ---- load data ----
    if not os.path.exists(PICKS_PATH):
        print(f"No picks file found at {PICKS_PATH} — nothing to export.", file=sys.stderr)
        sys.exit(0)

    with open(PICKS_PATH) as f:
        picks_data = json.load(f)

    scores_data = {}
    if os.path.exists(SCORES_PATH):
        with open(SCORES_PATH) as f:
            scores_data = json.load(f)

    participants = picks_data.get("participants", [])
    all_picks    = picks_data.get("picks", {})
    results      = scores_data.get("results", {})    # not in scores.json by default; placeholder

    def get_pick(pid, sid):
        raw = all_picks.get(pid, {}).get(sid)
        if not raw:           return {"winner": None, "games": None}
        if isinstance(raw, str): return {"winner": raw, "games": None}
        return {"winner": raw.get("winner"), "games": raw.get("games")}

    def actual_games(sid):
        rec = scores_data.get("records", {})
        # find the record key that matches this series (we store by abbr pair)
        for k, v in rec.items():
            wins = list(v.values())
            total = sum(wins)
            if total >= 4 and max(wins) >= 4:
                # we can't easily map sid→abbr pair here without full TEAMS dict
                pass
        return None  # game count from API not critical for Excel

    def score_for(pid):
        pts = 0
        for s in SERIES:
            actual = results.get(s["id"])
            pick   = get_pick(pid, s["id"])
            if actual and pick["winner"] and actual == pick["winner"]:
                pts += ROUND_POINTS[s["r"]]
        return pts

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════
    # Summary sheet
    # ════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    hdr = ["Round", "Series"] + [p["name"] for p in participants] + ["Actual Winner"]
    for ci, h in enumerate(hdr, 1):
        cell(ws, 1, ci, h, fill=HDR_FILL, font=HDR_FONT, align="center")

    row = 2
    for s in SERIES:
        sid = s["id"]
        cell(ws, row, 1, ROUND_NAMES[s["r"]], fill=GREY_FILL)
        cell(ws, row, 2, sid, fill=GREY_FILL)
        actual = results.get(sid)
        for ci, p in enumerate(participants, 3):
            pick = get_pick(p["id"], sid)
            val  = pick["winner"] or ""
            if pick["games"]: val += f" in {pick['games']}"
            is_ok  = actual and pick["winner"] and actual == pick["winner"]
            is_bad = actual and pick["winner"] and actual != pick["winner"]
            fill = GREEN_FILL if is_ok else (RED_FILL if is_bad else GREY_FILL)
            cell(ws, row, ci, val, fill=fill)
        cell(ws, row, len(hdr), actual or "—", fill=GREY_FILL, bold=bool(actual))
        row += 1

    # Totals row
    cell(ws, row, 1, "TOTAL", fill=HDR_FILL, font=HDR_FONT)
    cell(ws, row, 2, "", fill=HDR_FILL)
    for ci, p in enumerate(participants, 3):
        cell(ws, row, ci, score_for(p["id"]), fill=HDR_FILL,
             font=Font(color="F97316", bold=True), align="center")
    cell(ws, row, len(hdr), "", fill=HDR_FILL)

    auto_width(ws)

    # ════════════════════════════════════════
    # One sheet per participant
    # ════════════════════════════════════════
    for p in participants:
        safe = p["name"].replace("/", "_").replace("\\", "_")[:31]
        ws2  = wb.create_sheet(safe)
        ws2.sheet_view.showGridLines = False

        hdrs2 = ["Round", "Series", "Your Pick", "Games", "Actual", "Points"]
        for ci, h in enumerate(hdrs2, 1):
            cell(ws2, 1, ci, h, fill=HDR_FILL, font=HDR_FONT, align="center")

        row2 = 2
        total_pts = 0
        for s in SERIES:
            sid    = s["id"]
            pick   = get_pick(p["id"], sid)
            actual = results.get(sid)
            pts    = 0
            if actual and pick["winner"] and actual == pick["winner"]:
                pts = ROUND_POINTS[s["r"]]
            total_pts += pts
            is_ok  = actual and pick["winner"] and actual == pick["winner"]
            is_bad = actual and pick["winner"] and actual != pick["winner"]
            fill   = GREEN_FILL if is_ok else (RED_FILL if is_bad else GREY_FILL)
            cell(ws2, row2, 1, ROUND_NAMES[s["r"]], fill=GREY_FILL)
            cell(ws2, row2, 2, sid, fill=GREY_FILL)
            cell(ws2, row2, 3, pick["winner"] or "—", fill=fill)
            cell(ws2, row2, 4, pick["games"] or "—", fill=fill, align="center")
            cell(ws2, row2, 5, actual or "—", fill=GREY_FILL)
            cell(ws2, row2, 6, pts if pts else "—", fill=fill, align="center")
            row2 += 1

        cell(ws2, row2, 1, "TOTAL", fill=HDR_FILL, font=HDR_FONT)
        for ci in range(2, 6):
            cell(ws2, row2, ci, "", fill=HDR_FILL)
        cell(ws2, row2, 6, total_pts, fill=HDR_FILL, font=Font(color="F97316", bold=True), align="center")
        auto_width(ws2)

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}  ({len(participants)} participants, {len(SERIES)} series)")


if __name__ == "__main__":
    main()
